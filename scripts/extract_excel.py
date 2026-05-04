#!/usr/bin/env python3
"""Extract candidate data from an SJMSOM PhD admissions Excel file.

Usage: extract_excel.py <path-to-xlsx>
Outputs JSON array to stdout. Exits non-zero on failure.
"""
import json
import re
import sys
from openpyxl import load_workbook


def normalize_categories_applied(s):
    """RA and RA/TA collapse to TA; standalone RA tokens in lists become TA.
    Mirrors normalize_categories_applied() in src/helpers.php — keep both in sync."""
    if s is None:
        return None
    s = str(s).strip()
    if not s:
        return ""
    compact = re.sub(r"\s+", "", s.upper())
    if compact in ("RA", "RA/TA", "TA/RA"):
        return "TA"
    return re.sub(r"\bRA\b", "TA", s, flags=re.IGNORECASE)


# Map of canonical field name -> list of header substrings (lowercased) that match
HEADER_ALIASES = {
    "serial_no": ["serial no", "sr. no", "s.no", "sl no"],
    "applicant_id": ["applicantid", "applicant id"],
    "dept_reg_no": ["dept reg. no", "dept reg no", "dept. reg. no", "department reg"],
    "name": ["name"],
    "email": ["email"],
    "gender": ["gender"],
    "birth_category": ["birth category"],
    "ews": ["ews"],
    "disabled": ["disabled"],
    "cfti": ["cfti"],
    "iit_btech": ["iit btech", "iit b.tech", "iit b tech"],
    "categories_applied": ["categoriesapplied", "categories applied"],
    "academic_record": ["academic record"],
    "qualifying_exam": ["qualifying exam"],
    "exam_status": ["exam status"],
    "qualifying_discipline": ["qualifying discipline"],
    "passing_year": ["passing year"],
    "percentage": ["percentage"],
    "original_percentage": ["original percentage"],
    "original_percentage_out_of": ["original percentage out of", "out of"],
    "cpi_grade": ["cpi", "grade"],
    "gate_score": ["gate score"],
    "gate_year": ["gate year"],
    "gate_regn": ["gate regn", "gate reg"],
    "work_experience": ["work experience"],
    "fellowship": ["fellowship"],
    "research_interest_selected": ["research interests (for selected", "research interest (for selected"],
    "research_interest_other": ["research interests (for other", "research interest (for other"],
    "remark": ["remark"],
}

# Order of preference - the more specific should come first so "original percentage"
# beats "percentage" for its column.
SPECIFIC_FIRST = [
    "original_percentage_out_of",
    "original_percentage",
    "research_interest_selected",
    "research_interest_other",
    "gate_score",
    "gate_year",
    "gate_regn",
]


def norm(s):
    if s is None:
        return ""
    return str(s).strip().lower()


def build_header_map(header_row):
    """Return dict: canonical_field -> column_index (0-based)."""
    headers = [norm(c) for c in header_row]
    mapping = {}
    taken = set()

    # Pass 1: specific headers first
    ordered_fields = SPECIFIC_FIRST + [f for f in HEADER_ALIASES if f not in SPECIFIC_FIRST]
    for field in ordered_fields:
        for idx, h in enumerate(headers):
            if idx in taken:
                continue
            for alias in HEADER_ALIASES[field]:
                if alias in h:
                    mapping[field] = idx
                    taken.add(idx)
                    break
            if field in mapping:
                break

    # Pass 2: fallback for generic "Research Interests" column when the sheet
    # doesn't split it into (for selected) / (for other) variants.
    if "research_interest_selected" not in mapping:
        for idx, h in enumerate(headers):
            if idx in taken:
                continue
            if "research interest" in h and "other" not in h:
                mapping["research_interest_selected"] = idx
                taken.add(idx)
                break

    return mapping


def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def main():
    if len(sys.argv) < 2:
        print("Usage: extract_excel.py <path-to-xlsx>", file=sys.stderr)
        sys.exit(2)
    path = sys.argv[1]
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Failed to open workbook: {e}"}), file=sys.stderr)
        sys.exit(3)

    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        print(json.dumps({"error": "Empty worksheet"}), file=sys.stderr)
        sys.exit(4)

    mapping = build_header_map(list(header))
    if "dept_reg_no" not in mapping or "name" not in mapping:
        print(json.dumps({
            "error": "Required columns not found: Dept Reg. No. and Name",
            "detected_headers": [norm(c) for c in header],
        }), file=sys.stderr)
        sys.exit(5)

    out = []
    for row in rows_iter:
        if row is None:
            continue
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        rec = {}
        for field, idx in mapping.items():
            val = row[idx] if idx < len(row) else None
            rec[field] = cell_str(val)
        if not rec.get("dept_reg_no") or not rec.get("name"):
            continue
        if "categories_applied" in rec:
            rec["categories_applied"] = normalize_categories_applied(rec["categories_applied"])
        out.append(rec)

    json.dump(out, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
