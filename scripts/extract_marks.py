#!/usr/bin/env python3
"""Extract written marks from Excel - expects Dept Reg No + Written Marks columns."""
import sys, json, re
try:
    from openpyxl import load_workbook
except ImportError:
    print(json.dumps({"error": "openpyxl not installed"}))
    sys.exit(1)

if len(sys.argv) < 2:
    print(json.dumps([]))
    sys.exit(0)

path = sys.argv[1]
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))
if not rows:
    print(json.dumps([]))
    sys.exit(0)

header = [str(c).strip().lower() if c else '' for c in rows[0]]

def find_col(aliases):
    for alias in aliases:
        for i, h in enumerate(header):
            if alias in h:
                return i
    return None

reg_col   = find_col(['dept reg', 'dept_reg', 'reg no', 'regn no', 'rmg'])
marks_col = find_col(['written marks', 'written_marks', 'marks', 'score'])

if reg_col is None or marks_col is None:
    sys.stderr.write(f"Could not find required columns. Headers: {header}\n")
    sys.exit(1)

results = []
for row in rows[1:]:
    dept = str(row[reg_col]).strip() if row[reg_col] is not None else ''
    mark_raw = row[marks_col]
    if not dept or dept.lower() in ('none', 'null', ''):
        continue
    try:
        mark = float(mark_raw) if mark_raw is not None else None
    except (ValueError, TypeError):
        mark = None
    if mark is None:
        continue
    results.append({'dept_reg_no': dept, 'written_marks': mark})

print(json.dumps(results))
