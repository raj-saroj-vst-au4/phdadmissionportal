#!/usr/bin/env python3
"""Extract international candidate data from an Excel file.

Expected columns: Sr. No., Student ID, Name, Email, Nationality, Research Area.

Usage: extract_excel_intl.py <path-to-xlsx>
Outputs JSON array to stdout. Exits non-zero on failure.
"""
import json
import sys
from openpyxl import load_workbook


HEADER_ALIASES = {
    "serial_no": ["sr. no", "sr no", "serial no", "s.no", "sl no"],
    "student_id": ["student id", "studentid", "applicant id", "applicantid"],
    "name": ["name"],
    "email": ["email"],
    "nationality": ["nationality", "country"],
    "research_area": ["research area", "research interest", "research"],
}


def norm(s):
    if s is None:
        return ""
    return str(s).strip().lower()


def build_header_map(header_row):
    headers = [norm(c) for c in header_row]
    mapping = {}
    taken = set()
    for field, aliases in HEADER_ALIASES.items():
        for idx, h in enumerate(headers):
            if idx in taken:
                continue
            for alias in aliases:
                if alias in h:
                    mapping[field] = idx
                    taken.add(idx)
                    break
            if field in mapping:
                break
    return mapping


def cell_str(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def main():
    if len(sys.argv) < 2:
        print("Usage: extract_excel_intl.py <path-to-xlsx>", file=sys.stderr)
        sys.exit(2)
    path = sys.argv[1]
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as e:
        print(json.dumps({"error": f"Failed to open workbook: {e}"}), file=sys.stderr)
        sys.exit(3)

    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        print(json.dumps({"error": "Empty worksheet"}), file=sys.stderr)
        sys.exit(4)

    mapping = build_header_map(list(header))
    if "student_id" not in mapping or "name" not in mapping:
        print(json.dumps({
            "error": "Required columns not found: Student ID and Name",
            "detected_headers": [norm(c) for c in header],
        }), file=sys.stderr)
        sys.exit(5)

    out = []
    for row in rows_iter:
        if row is None:
            continue
        if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
            continue
        rec = {}
        for field, idx in mapping.items():
            val = row[idx] if idx < len(row) else None
            rec[field] = cell_str(val)
        if not rec.get("student_id") or not rec.get("name"):
            continue
        out.append(rec)

    json.dump(out, sys.stdout, ensure_ascii=False)


if __name__ == "__main__":
    main()
